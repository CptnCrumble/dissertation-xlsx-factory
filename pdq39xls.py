import json
import os
import urllib.request as req
from openpyxl import Workbook

def pdq39xls(pid):
    ip = os.getenv('HOST_IP')
    port = os.getenv('PG_ADAPTOR_PORT')
    url = f'http://{ip}:{port}/pdq39data/{pid}'
    jdata = req.urlopen(url).read()
    # data is a list of dictionaries, each dic is a form entry from the db
    data = json.loads(jdata)    

    # creating new Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Example"

    # Build top row
    ws['A1'] = "Patient"
    ws['B1'] = "Assessment"
    ws['C1'] = "Date"
    ws['D1'] = "(1) leisure"
    ws['E1'] = "(2) DIY"
    ws['F1'] = "(3) bags"
    ws['G1'] = "(4) 0.5 mile"
    ws['H1'] = "(5) 100 yds"
    ws['I1'] = "(6) house"
    ws['J1'] = "(7) public"
    ws['K1'] = "(8) else"
    ws['L1'] = "(9) fright"
    ws['M1'] = "(10) confined"
    ws['N1'] = "(11) wash"
    ws['O1'] = "(12) dress"
    ws['P1'] = "(13) buttons"
    ws['Q1'] = "(14) write"
    ws['R1'] = "(15) food"
    ws['S1'] = "(16) drink"
    ws['T1'] = "(17) depressed"
    ws['U1'] = "(18) lonely"
    ws['V1'] = "(19) weepy"
    ws['W1'] = "(20) angry"
    ws['X1'] = "(21) anxious"
    ws['Y1'] = "(22) future"
    ws['Z1'] = "(23) conceal"
    ws['AA1'] = "(24) avoid"
    ws['AB1'] = "(25) embarrassed"
    ws['AC1'] = "(26) reactions"
    ws['AD1'] = "(27) relationships"
    ws['AE1'] = "(28) support"
    ws['AF1'] = "(29) family"
    ws['AG1'] = "(30) asleep"
    ws['AH1'] = "(31) concern"
    ws['AI1'] = "(32) memory"
    ws['AJ1'] = "(33) dreams"
    ws['AK1'] = "(34) speech"
    ws['AL1'] = "(35) communicate"
    ws['AM1'] = "(36) ignored"
    ws['AN1'] = "(37) cramps"
    ws['AO1'] = "(38) pain"
    ws['AP1'] = "(39) hot"
    ws['AS1'] = "PDQ-SI"
    ws['AT1'] = "Mob"
    ws['AU1'] = "ADL"
    ws['AV1'] = "Emot"
    ws['AW1'] = "Stigma"
    ws['AX1'] = "Soc Sup"
    ws['AY1'] = "Cog"
    ws['AZ1'] = "Comm"
    ws['BA1'] = "Discom"


    # Write in some data
    active_row = 2
    for record in data:        
        ws[f'A{active_row}'].value = record['Pid']
        ws[f'B{active_row}'].value = record['AssessmentNumber']
        ws[f'C{active_row}'].value = record['AssessmentDate'][0:10]
        ws[f'D{active_row}'].value = record['Pdq1']
        ws[f'E{active_row}'].value = record['Pdq2']
        ws[f'F{active_row}'].value = record['Pdq3']
        ws[f'G{active_row}'].value = record['Pdq4']
        ws[f'H{active_row}'].value = record['Pdq5']
        ws[f'I{active_row}'].value = record['Pdq6']
        ws[f'J{active_row}'].value = record['Pdq7']
        ws[f'K{active_row}'].value = record['Pdq8']
        ws[f'L{active_row}'].value = record['Pdq9']
        ws[f'M{active_row}'].value = record['Pdq10']
        ws[f'N{active_row}'].value = record['Pdq11']
        ws[f'O{active_row}'].value = record['Pdq12']
        ws[f'P{active_row}'].value = record['Pdq13']
        ws[f'Q{active_row}'].value = record['Pdq14']
        ws[f'R{active_row}'].value = record['Pdq15']
        ws[f'S{active_row}'].value = record['Pdq16']
        ws[f'T{active_row}'].value = record['Pdq17']
        ws[f'U{active_row}'].value = record['Pdq18']
        ws[f'V{active_row}'].value = record['Pdq19']
        ws[f'W{active_row}'].value = record['Pdq20']
        ws[f'X{active_row}'].value = record['Pdq21']
        ws[f'Y{active_row}'].value = record['Pdq22']
        ws[f'Z{active_row}'].value = record['Pdq23']
        ws[f'AA{active_row}'].value = record['Pdq24']
        ws[f'AB{active_row}'].value = record['Pdq25']
        ws[f'AC{active_row}'].value = record['Pdq26']
        ws[f'AD{active_row}'].value = record['Pdq27']
        ws[f'AE{active_row}'].value = record['Pdq28']
        ws[f'AF{active_row}'].value = record['Pdq29']
        ws[f'AG{active_row}'].value = record['Pdq30']
        ws[f'AH{active_row}'].value = record['Pdq31']
        ws[f'AI{active_row}'].value = record['Pdq32']
        ws[f'AJ{active_row}'].value = record['Pdq33']
        ws[f'AK{active_row}'].value = record['Pdq34']
        ws[f'AL{active_row}'].value = record['Pdq35']
        ws[f'AM{active_row}'].value = record['Pdq36']
        ws[f'AN{active_row}'].value = record['Pdq37']
        ws[f'AO{active_row}'].value = record['Pdq38']
        ws[f'AP{active_row}'].value = record['Pdq39']
        ws[f'AS{active_row}'] = f'=(AT{active_row}+AU{active_row}+AV{active_row}+AW{active_row}+AX{active_row}+AY{active_row}+AZ{active_row}+BA{active_row})/8'
        ws[f'AT{active_row}'] = f'=(D{active_row}+E{active_row}+F{active_row}+G{active_row}+H{active_row}+I{active_row}+J{active_row}+K{active_row}+L{active_row}+M{active_row})/0.4'
        ws[f'AU{active_row}'] = f'=(N{active_row}+O{active_row}+P{active_row}+Q{active_row}+R{active_row}+S{active_row})/0.24'
        ws[f'AV{active_row}'] = f'=(T{active_row}+U{active_row}+V{active_row}+W{active_row}+X{active_row}+Y{active_row})/0.24'
        ws[f'AW{active_row}'] = f'=(Z{active_row}+AA{active_row}+AB{active_row}+AC{active_row})/0.16'
        ws[f'AX{active_row}'] = f'=(AD{active_row}+AE{active_row}+AF{active_row})/0.12'
        ws[f'AY{active_row}'] = f'=(AG{active_row}+AH{active_row}+AI{active_row}+AJ{active_row})/0.16'
        ws[f'AZ{active_row}'] = f'=(AK{active_row}+AL{active_row}+AM{active_row})/0.12'
        ws[f'BA{active_row}'] = f'=(AN{active_row}+AO{active_row}+AP{active_row})/0.12'
        active_row += 1

    return wb

